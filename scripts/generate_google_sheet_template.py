from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "templates" / "google-sheet-template.xlsx"


def style_sheet(ws, headers, rows):
    ws.append(headers)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill(fill_type="solid", fgColor="0F6C5B")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(headers))
    last_row = len(rows) + 1
    table = Table(displayName=f"{ws.title}Table", ref=f"A1:{last_col}{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    ws.add_table(table)


def main():
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    stores = workbook.active
    stores.title = "Stores"
    level_targets = workbook.create_sheet("LevelTargets")
    results = workbook.create_sheet("DailyResults")
    readme = workbook.create_sheet("README")

    store_headers = [
        "code",
        "password",
        "name",
        "area",
        "target_prepaid_new_line",
        "target_prepaid_portabilidad",
        "target_postpaid_new_line",
        "target_postpaid_portabilidad",
        "reward_prepaid_new_line",
        "reward_prepaid_portabilidad",
        "reward_postpaid_new_line",
        "reward_postpaid_portabilidad"
    ]
    store_rows = [
        ["CUSPS0001", "123456", "Punto San Isidro", "Lima", 30, 12, 8, 5, 50, 60, 70, 80],
        ["CUSPS0002", "abc789", "Punto Miraflores", "Lima", 28, 10, 9, 6, 50, 60, 70, 80],
        ["CUSPS0003", "pdv2026", "Punto Arequipa Centro", "Arequipa", 22, 8, 7, 4, 50, 60, 70, 80]
    ]

    level_target_headers = [
        "store_code",
        "level_code",
        "level_order",
        "level_name",
        "reward",
        "target_prepaid_new_line",
        "target_prepaid_portabilidad",
        "target_postpaid_new_line",
        "target_postpaid_portabilidad"
    ]
    level_target_rows = [
        ["CUSPS0001", "M1", 1, "Mức 1", 50, 1, 2, 1, 4],
        ["CUSPS0001", "M2", 2, "Mức 2", 100, 3, 4, 2, 6],
        ["CUSPS0001", "M3", 3, "Mức 3", 150, 5, 6, 3, 8],
        ["CUSPS0002", "M1", 1, "Mức 1", 50, 1, 1, 1, 3],
        ["CUSPS0002", "M2", 2, "Mức 2", 100, 2, 3, 2, 5],
        ["CUSPS0002", "M3", 3, "Mức 3", 150, 4, 5, 3, 7],
        ["CUSPS0003", "M1", 1, "Mức 1", 50, 1, 1, 1, 2],
        ["CUSPS0003", "M2", 2, "Mức 2", 100, 2, 2, 2, 4],
        ["CUSPS0003", "M3", 3, "Mức 3", 150, 3, 4, 3, 5]
    ]

    result_headers = [
        "date",
        "store_code",
        "prepaid_new_line",
        "prepaid_portabilidad",
        "postpaid_new_line",
        "postpaid_portabilidad",
        "note"
    ]
    result_rows = [
        ["2026-04-06", "CUSPS0001", 1, 1, 0, 2, "Inicio de semana"],
        ["2026-04-07", "CUSPS0001", 1, 1, 1, 1, "Buen avance"],
        ["2026-04-08", "CUSPS0001", 2, 1, 1, 2, "Dia actual"],
        ["2026-04-06", "CUSPS0002", 1, 0, 1, 1, "Base diaria"],
        ["2026-04-07", "CUSPS0002", 1, 1, 0, 2, "Porta estable"],
        ["2026-04-08", "CUSPS0002", 2, 1, 1, 1, "Dia actual"],
        ["2026-04-06", "CUSPS0003", 0, 1, 1, 0, "Apertura"],
        ["2026-04-07", "CUSPS0003", 1, 0, 1, 1, "Dia previo"],
        ["2026-04-08", "CUSPS0003", 1, 1, 1, 1, "Dia actual"]
    ]

    style_sheet(stores, store_headers, store_rows)
    style_sheet(level_targets, level_target_headers, level_target_rows)
    style_sheet(results, result_headers, result_rows)

    for column, width in {
        "A": 14,
        "B": 14,
        "C": 24,
        "D": 16,
        "E": 24,
        "F": 28,
        "G": 24,
        "H": 28,
        "I": 18,
        "J": 20,
        "K": 18,
        "L": 20
    }.items():
        stores.column_dimensions[column].width = width

    for column, width in {
        "A": 14,
        "B": 14,
        "C": 12,
        "D": 16,
        "E": 12,
        "F": 24,
        "G": 28,
        "H": 24,
        "I": 28
    }.items():
        level_targets.column_dimensions[column].width = width

    for column, width in {
        "A": 14,
        "B": 14,
        "C": 20,
        "D": 24,
        "E": 20,
        "F": 24,
        "G": 22
    }.items():
        results.column_dimensions[column].width = width

    readme["A1"] = "Huong dan cap nhat du lieu v4"
    readme["A1"].font = Font(bold=True, size=14)
    readme["A3"] = "1. Sheet Stores: khai bao user, password, chi tieu tong va tien thuong cua 4 loai thue bao."
    readme["A4"] = "2. Sheet LevelTargets: moi dong la 1 muc chi tieu cua 1 diem ban, kem reward va 4 chi tieu con."
    readme["A5"] = "3. Muon them muc moi trong tuong lai, chi can them 1 dong moi vao LevelTargets."
    readme["A6"] = "4. Sheet DailyResults: moi dong la ket qua 1 diem ban trong 1 ngay."
    readme["A7"] = "5. App chi cong nhan dat muc khi ca 4 loai thue bao cua muc do deu dat."
    readme["A9"] = "Tai khoan test:"
    readme["A10"] = "CUSPS0001 / 123456"
    readme["A11"] = "CUSPS0002 / abc789"
    readme["A12"] = "CUSPS0003 / pdv2026"
    readme.column_dimensions["A"].width = 100

    workbook.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
